"""Inspect sheet RESUMEN of SEMANA 6 INT 40 to map structure."""
from openpyxl import load_workbook
from pathlib import Path

base = Path(r"C:\Users\agath\Downloads\Casa Nogal")
semana = next(base.glob("SEMANA*INT 40*.xlsx"))
wb = load_workbook(semana, data_only=True, read_only=False)
ws = wb["RESUMEN"]

print(f"Sheet RESUMEN: {ws.max_row} rows x {ws.max_column} cols\n")

# Pre-load to memory
MAX_R, MAX_C = 80, 65
grid = [[None] * (MAX_C + 1) for _ in range(MAX_R + 1)]
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_R, max_col=MAX_C, values_only=True), 1):
    for ci, v in enumerate(row, 1):
        grid[ri][ci] = v

for r in range(1, 60):
    cells = [(c, grid[r][c]) for c in range(1, MAX_C + 1) if grid[r][c] not in (None, "")]
    if cells:
        compact = " | ".join(f"{c}={v}" for c, v in cells[:15])
        more = f" ... +{len(cells) - 15}" if len(cells) > 15 else ""
        print(f"R{r}: {compact}{more}")

wb.close()
