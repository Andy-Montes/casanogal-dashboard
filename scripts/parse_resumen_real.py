"""Parse sheet RESUMEN of SEMANA 6 INT 40 into int40_real.json.

Estructura del RESUMEN (mapeada):
- Cada niño ocupa ~10 filas. Su fila principal (header) tiene el nombre en col 2
  y la grilla de horario en cols 3..55.
- Cols por día: 8 franjas cada uno, con 1 col de gap.
  Día 0: cols 3-10, Día 1: cols 12-19, Día 2: cols 21-28,
  Día 3: cols 30-37, Día 4: cols 39-46, Día 5: cols 48-55.
- Las sub-filas "SEM 1..6" son meta-info (observaciones, talleres, coaching),
  NO las sesiones individuales niño-terapeuta. Ignoradas para el horario real.
- Por lo tanto el RESUMEN solo tiene 1 semana de patrón. Las 6 semanas del
  intensivo repiten ese patrón.

Output: data/intensivos/int40_real.json con la misma shape que el output del
scheduler:
{
  "fuente": "horario real de Trini (sheet RESUMEN del SEMANA 6 INT 40)",
  "semanas": [
    { "grid": [ [sigla|null * 48 slots] * 6 niños ], "sesionesPlanificadas": N }
    ... x6
  ]
}
"""
import json
from openpyxl import load_workbook
from pathlib import Path

base = Path(r"C:\Users\agath\Downloads\Casa Nogal")
semana = next(base.glob("SEMANA*INT 40*.xlsx"))
wb = load_workbook(semana, data_only=True, read_only=False)
ws = wb["RESUMEN"]

# Pre-cargar a memoria
MAX_R, MAX_C = 80, 70
grid = [[None] * (MAX_C + 1) for _ in range(MAX_R + 1)]
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_R, max_col=MAX_C, values_only=True), 1):
    for ci, v in enumerate(row, 1):
        grid[ri][ci] = v

# Mapeo niño → fila header
NIÑOS_FILAS = [
    ("LEON",     4),
    ("SANTI",   14),
    ("BALTAZAR", 24),
    ("FIORELLA", 34),
    ("AMELIA",   44),
    ("SOFIA",    54),
]
# Verificar que el nombre en col 2 corresponde
for nombre_corto, fila in NIÑOS_FILAS:
    real = str(grid[fila][2] or "").strip().upper()
    real_norm = (real.replace("Á","A").replace("É","E").replace("Í","I")
                     .replace("Ó","O").replace("Ú","U").split()[0] if real else "")
    if real_norm != nombre_corto:
        print(f"⚠ Fila {fila} esperaba {nombre_corto}, tiene '{real}'")

# Cols por día: día d → primera col = 3 + d*9
DIAS = 6
FRANJAS = 8

def col_de(dia, franja):
    return 3 + dia * 9 + franja

catalogo_path = Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\catalogo.json")
SIGLAS_VALIDAS = set(json.loads(catalogo_path.read_text(encoding="utf-8"))["terapeutas"].keys())
# KIDS aparece en celdas grupales — siempre se mapea a GP (terapeuta a cargo)
SIGLAS_VALIDAS_PLUS_KIDS = SIGLAS_VALIDAS | {"KIDS"}

def normalizar_sigla(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s or s == "_" or s == "-":
        return None
    if s == "KIDS":
        return "GP"  # KIDS grupal → terapeuta GP, así casa con scheduler.kidsSlots
    if s in SIGLAS_VALIDAS:
        return s
    return None  # nombres de niños, "PAPÁS X", "OBS X", etc. → ignorar

# Construir 1 semana × 6 niños × 48 slots
semana_grid = []
for nombre_corto, fila in NIÑOS_FILAS:
    slots = []
    for d in range(DIAS):
        for f in range(FRANJAS):
            slots.append(normalizar_sigla(grid[fila][col_de(d, f)]))
    semana_grid.append(slots)

# Resumen de validación
print("\nSesiones por niño (1 semana):")
for i, (nombre, _) in enumerate(NIÑOS_FILAS):
    n = sum(1 for s in semana_grid[i] if s)
    siglas_unicas = set(s for s in semana_grid[i] if s)
    print(f"  {nombre:10s}: {n:>2} slots ocupados · siglas únicas: {sorted(siglas_unicas)}")

# Replicar para 6 semanas (mismo patrón base)
semanas = []
sesPlan = sum(sum(1 for s in row if s) for row in semana_grid)
for si in range(6):
    semanas.append({
        "grid": [list(row) for row in semana_grid],  # clones
        "sesionesPlanificadas": sesPlan,
    })

out = {
    "fuente": "horario real de Trini · sheet RESUMEN del SEMANA 6 INT 40 (1 semana base, replicada x6)",
    "ok": True,
    "semanas": semanas,
}

dest = Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\int40_real.json")
dest.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nWrote {dest}")
print(f"Total slots ocupados por semana: {sesPlan}")
wb.close()
