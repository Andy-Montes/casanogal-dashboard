"""Derive sala capacity from INT 40 real schedule.

SALAS sheet structure:
  R2: date in col 3, 10, 17, 24, 31, 38 (6 días, 7 cols each)
  R3: franja text in each col (cols 3-9 = día 1, etc.)
  R4: niño 1 nombre (col 2) + salas en cols 3..44 (7×6=42)
  R5: terapeuta sigla (paired)
  R7,R10,R13,R16,R19,R22: siguientes niños
"""
import json
from collections import defaultdict, Counter
from openpyxl import load_workbook
from pathlib import Path

base = Path(r"C:\Users\agath\Downloads\Casa Nogal")
semana = next(base.glob("SEMANA*INT 40*.xlsx"))
wb = load_workbook(semana, data_only=True, read_only=True)
ws = wb["SALAS"]

# Filas de los niños del intensivo
NINO_ROWS = [4, 7, 10, 13, 16, 19, 22]
# Cols por día (7 franjas/día, 6 días)
DAYS = 6
FRANJAS_PER_DAY = 7
COL_START = 3

# slot → counter{sala: count}
slot_sala = defaultdict(Counter)

for row in NINO_ROWS:
    nombre = ws.cell(row, 2).value
    if not nombre:
        continue
    for d in range(DAYS):
        for f in range(FRANJAS_PER_DAY):
            col = COL_START + d * FRANJAS_PER_DAY + f
            sala = ws.cell(row, col).value
            if sala and str(sala).strip():
                s = str(sala).strip()
                slot_sala[(d, f)][s] += 1

# Capacidad observada por sala = max sobre todos los slots
cap = defaultdict(int)
for slot, counts in slot_sala.items():
    for sala, n in counts.items():
        if n > cap[sala]:
            cap[sala] = n

print("Capacidad observada por sala (max sesiones simultáneas en INT 40):")
for sala, c in sorted(cap.items(), key=lambda x: -x[1]):
    print(f"  {sala:12s} → {c}")

# Persistir como JSON
out = Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\salas_capacidad.json")
out.write_text(json.dumps(dict(cap), ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nWrote {out}")
wb.close()
