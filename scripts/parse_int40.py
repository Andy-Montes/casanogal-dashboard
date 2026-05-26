"""Parse INT 40 sheet into clean intensivos/int40.json.

Output structure:
{
  "id": "INT 40",
  "fecha_inicio": "2026-04-20",
  "fecha_fin": "2026-05-29",
  "semanas": 6,
  "niños": [
    {
      "nombre": "LEON",
      "encargado": "Paula Andrades",
      "kids_semanal": 5,
      "total_ses_semanal": 30,
      "asignaciones": [
        { "disciplina": "TO",   "rol": "TUTOR", "sigla": "KRA", "sesiones": 4 },
        { "disciplina": "TO",   "rol": "COT",   "sigla": "PT",  "sesiones": 3 },
        { "disciplina": "FONO", "rol": "TUTOR", "sigla": "NP",  "sesiones": 4 },
        ...
      ]
    },
    ...
  ]
}
"""
import json, re
from openpyxl import load_workbook
from pathlib import Path

base = Path(r"C:\Users\agath\Downloads\Casa Nogal")
equipos = next(base.glob("Equipos*Intensivos*.xlsx"))
wb = load_workbook(equipos, data_only=True, read_only=False)
ws = wb["INT 40"]

# Bloques por disciplina segun R3/R4: (disciplina, [(rol, col), ...])
BLOQUES = [
    ("TO",   [("TUTOR", 4),  ("COT", 5)]),
    ("FONO", [("TUTOR", 6),  ("COT", 7)]),
    ("COG",  [("TUTOR", 8),  ("COT", 9),  ("AULA", 10)]),
    ("KINE", [("TUTOR", 11), ("COT", 12), ("AULA", 13)]),
    ("PSI",  [("TUTOR", 14), ("COT", 15), ("PAPAS", 16)]),
]
COL_KIDS = 17
COL_TOTAL = 18
COL_NOMBRE = 2
COL_ENCARGADO = 3

def parse_sigla_field(v):
    """Devuelve lista [(sigla, sesiones_override_o_None), ...]."""
    if v is None or str(v).strip() == "":
        return []
    s = str(v).strip()
    parts = re.split(r"[/,]", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(r"^([A-ZÁÉÍÓÚÑa-z]+)\s*(\d+)?$", p)
        if m:
            out.append((m.group(1).upper(), int(m.group(2)) if m.group(2) else None))
        else:
            out.append((p.upper(), None))
    return out

def parse_count(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    m = re.match(r"^(\d+)", s)
    return int(m.group(1)) if m else None

ninos = []
# Recorrer pares de filas desde R5 hasta donde haya datos
row = 5
while row <= ws.max_row:
    name_cell = ws.cell(row, COL_NOMBRE).value
    if not name_cell or str(name_cell).strip() == "":
        break
    nombre = str(name_cell).strip()
    encargado = ws.cell(row, COL_ENCARGADO).value
    encargado = str(encargado).strip() if encargado else ""

    asignaciones = []
    for disciplina, roles in BLOQUES:
        for rol, col in roles:
            sigla_v = ws.cell(row, col).value
            count_v = ws.cell(row + 1, col).value
            siglas = parse_sigla_field(sigla_v)
            total_count = parse_count(count_v)
            if not siglas:
                continue
            if len(siglas) == 1:
                sigla, override = siglas[0]
                ses = override if override is not None else total_count
                if ses:
                    asignaciones.append({
                        "disciplina": disciplina,
                        "rol": rol,
                        "sigla": sigla,
                        "sesiones": ses,
                    })
            else:
                # Co-tutores divididos: "ANT2/KRA1"
                for sigla, override in siglas:
                    if override:
                        asignaciones.append({
                            "disciplina": disciplina,
                            "rol": rol,
                            "sigla": sigla,
                            "sesiones": override,
                        })

    kids = parse_count(ws.cell(row + 1, COL_KIDS).value) or 0
    total = parse_count(ws.cell(row + 1, COL_TOTAL).value) or 0

    ninos.append({
        "nombre": nombre,
        "encargado": encargado,
        "kids_semanal": kids,
        "total_ses_semanal": total,
        "asignaciones": asignaciones,
    })
    row += 2

# Fechas del header
fecha_ini = ws.cell(1, 2).value
fecha_fin = ws.cell(1, 3).value
fmt = lambda d: d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)

result = {
    "id": "INT 40",
    "fecha_inicio": fmt(fecha_ini),
    "fecha_fin": fmt(fecha_fin),
    "semanas": 6,
    "niños": ninos,
}

out = Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\int40.json")
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out}")
print(f"  Niños: {len(ninos)}")
for n in ninos:
    sum_ses = sum(a["sesiones"] for a in n["asignaciones"])
    print(f"    {n['nombre']:12} | individuales={sum_ses:>3} | kids={n['kids_semanal']} | total declarado={n['total_ses_semanal']} | asign={len(n['asignaciones'])}")

wb.close()
