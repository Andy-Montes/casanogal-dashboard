"""Parse availability per therapist from PROFESIONALES sheet (fast: bulk-load).

Layout: bandas de 10 filas (sigla+nombre, headers HORA/LUN..SÁB, 8 franjas).
7 terapeutas por banda en cols 2, 10, 18, 26, 34, 42, 50.

Output: data/intensivos/disponibilidad.json
  { SIGLA: { lun: [bool*8], mar: [bool*8], mie:..., jue:..., vie:..., sab:... } }
"""
import json
from openpyxl import load_workbook
from pathlib import Path

base = Path(r"C:\Users\agath\Downloads\Casa Nogal")
semana = next(base.glob("SEMANA*INT 40*.xlsx"))
wb = load_workbook(semana, data_only=True, read_only=False)
ws = wb["PROFESIONALES"]

catalogo = json.loads(Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\catalogo.json").read_text(encoding="utf-8"))
SIGLAS = set(catalogo["terapeutas"].keys())

intensivo = json.loads(Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\int40.json").read_text(encoding="utf-8"))
NINOS_INT = {n["nombre"].upper().split()[0] for n in intensivo["niños"]}  # {LEON, SANTI, ...}

def es_nino_intensivo(s):
    s = s.upper().replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U")
    primer = s.split()[0] if s.split() else ""
    return primer in NINOS_INT

def slot_disponible(v):
    if v is None:
        return True
    s = str(v).strip()
    if not s or s == "-":
        return True
    return es_nino_intensivo(s)

# Cargar 200 filas × 60 cols a memoria
MAX_R, MAX_C = 200, 60
grid = [[None] * (MAX_C + 1) for _ in range(MAX_R + 1)]
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_R, max_col=MAX_C, values_only=True), 1):
    for ci, v in enumerate(row, 1):
        grid[ri][ci] = v

DIAS = ["lun", "mar", "mie", "jue", "vie", "sab"]
FRANJAS_N = 8
COLS_TER = [2, 10, 18, 26, 34, 42, 50]

disponibilidad = {}

row = 1
while row <= MAX_R - 9:
    # ¿Es fila de siglas? buscar en COLS_TER
    siglas_en_banda = []
    for col in COLS_TER:
        v = grid[row][col]
        if v and str(v).strip() in SIGLAS:
            siglas_en_banda.append((col, str(v).strip()))
    if not siglas_en_banda:
        row += 1
        continue
    # Validar header HORA en fila siguiente
    h = grid[row + 1][2]
    if not h or str(h).strip().upper() != "HORA":
        row += 1
        continue

    for col, sigla in siglas_en_banda:
        slots = {d: [False] * FRANJAS_N for d in DIAS}
        for f in range(FRANJAS_N):
            for d_idx, d in enumerate(DIAS):
                cell = grid[row + 2 + f][col + 1 + d_idx]
                slots[d][f] = slot_disponible(cell)
        disponibilidad[sigla] = slots

    row += 10

out = Path(r"C:\Users\agath\Projects\dashboards\casanogal\data\intensivos\disponibilidad.json")
out.write_text(json.dumps(disponibilidad, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out}")
print(f"Terapeutas con disponibilidad: {len(disponibilidad)}/{len(SIGLAS)}")
faltantes = SIGLAS - set(disponibilidad.keys())
if faltantes:
    print(f"Sin datos: {sorted(faltantes)}")

print("\nSlots libres por terapeuta (max 48):")
for sigla in sorted(disponibilidad.keys()):
    total = sum(sum(s) for s in disponibilidad[sigla].values())
    print(f"  {sigla:5s} → {total:>2}/48")

wb.close()
